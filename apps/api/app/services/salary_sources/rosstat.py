from __future__ import annotations

import re
from io import BytesIO

import httpx

from app.services.salary_sources.common import USER_AGENT, normalize_text
from app.services.salary_sources.models import SalarySourceResult


ROSSTAT_PAGE = "https://rosstat.gov.ru/labor_market_employment_salaries"


def fetch_rosstat_region_wage(region: str, year: int, role: str = "") -> SalarySourceResult:
    try:
        response = httpx.get(ROSSTAT_PAGE, headers={"User-Agent": USER_AGENT}, timeout=10.0, follow_redirects=True)
    except httpx.TimeoutException:
        return _result("unavailable", role, region, year, notes="Официальный источник не был доступен при проверке.")
    except httpx.HTTPError as exc:
        return _result("unavailable", role, region, year, notes=f"Официальный источник не был доступен при проверке: {exc.__class__.__name__}.")

    if response.status_code in {403, 429}:
        return _result("blocked", role, region, year, source_url=ROSSTAT_PAGE, notes="Росстат ограничил доступ к странице при проверке.")
    if response.status_code >= 400:
        return _result("unavailable", role, region, year, source_url=ROSSTAT_PAGE, notes="Официальный источник не был доступен при проверке.")

    xlsx_url = _find_salary_xlsx_url(response.text)
    if not xlsx_url:
        return _result(
            "not_implemented",
            role,
            region,
            year,
            source_url=ROSSTAT_PAGE,
            notes="На странице Росстата не найден устойчивый XLSX с зарплатами по субъектам РФ.",
        )

    return _parse_rosstat_xlsx_url(xlsx_url, region, year, role)


def _find_salary_xlsx_url(html: str) -> str | None:
    candidates = re.findall(r"https?://[^\"']+?\.xlsx", html or "", flags=re.IGNORECASE)
    for url in candidates:
        lowered = url.lower()
        if "zar" in lowered or "salary" in lowered or "trud" in lowered:
            return url
    return candidates[0] if candidates else None


def _parse_rosstat_xlsx_url(url: str, region: str, year: int, role: str = "") -> SalarySourceResult:
    try:
        response = httpx.get(url, headers={"User-Agent": USER_AGENT}, timeout=20.0, follow_redirects=True)
    except httpx.HTTPError:
        return _result("unavailable", role, region, year, source_url=url, notes="XLSX Росстата не был доступен при проверке.")
    if response.status_code >= 400:
        return _result("unavailable", role, region, year, source_url=url, notes="XLSX Росстата не был доступен при проверке.")

    try:
        from openpyxl import load_workbook
    except Exception:
        return _result("not_implemented", role, region, year, source_url=url, notes="Для парсинга XLSX Росстата нужен openpyxl.")

    try:
        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    except Exception:
        return _result("parse_error", role, region, year, source_url=url, notes="Не удалось прочитать XLSX Росстата.")

    region_key = normalize_text(region).lower()
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        year_columns = [idx for row in rows[:20] for idx, cell in enumerate(row or []) if str(cell).strip() == str(year)]
        for row in rows:
            if not row:
                continue
            row_text = " ".join(str(cell) for cell in row if cell is not None).lower()
            if region_key not in row_text:
                continue
            for idx in year_columns:
                if idx < len(row) and isinstance(row[idx], (int, float)):
                    return SalarySourceResult(
                        source="rosstat",
                        status="ok",
                        query_role=role,
                        matched_role=None,
                        region=region,
                        year=year,
                        salary_value=int(round(row[idx])),
                        salary_type="official_region_mean",
                        source_url=url,
                        notes="Официальный региональный ориентир, не зарплата по конкретной должности.",
                    )
    return _result("no_data", role, region, year, source_url=url, notes="В XLSX Росстата не найден показатель для региона/года.")


def _result(status: str, role: str, region: str, year: int, source_url: str | None = ROSSTAT_PAGE, notes: str | None = None) -> SalarySourceResult:
    return SalarySourceResult(
        source="rosstat",
        status=status,
        query_role=role,
        matched_role=None,
        region=region,
        year=year,
        source_url=source_url,
        notes=notes,
    )

